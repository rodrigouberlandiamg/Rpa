from selenium import webdriver
from selenium.webdriver import chrome
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.by import By
from time import sleep
import pyautogui
from openpyxl import load_workbook
import threading
from multiprocessing import Process
import concurrent

class ChromeAuto:
    def __init__(self):
        self.driver_path = r"c:\path\to\chromedriver.exe"
        self.options = webdriver.ChromeOptions()
        self.chrome = webdriver.Chrome(executable_path=self.driver_path)

    def acessa(self,site):
        self.chrome.get(site)
        self.chrome.maximize_window()

    def sair(self):
        self.chrome.quit()

    def inicia_contagem(self):
        self.chrome.find_element_by_xpath('/html/body/app-root/div[2]/app-rpa1/div/div[1]/div[6]/button').click()

    def preenche_dados(self,lastname,address,email,firstname,phone,company,rolecompany):
        with concurrent.futures.ProcessPoolExecutor() as executor:
            executor.map([
        self.chrome.find_element_by_xpath('//input[@ng-reflect-name="labelLastName"]').send_keys(lastname),
        self.chrome.find_element_by_xpath('//input[@ng-reflect-name="labelAddress"]').send_keys(address),
        self.chrome.find_element_by_xpath('//input[@ng-reflect-name="labelEmail"]').send_keys(email),
        self.chrome.find_element_by_xpath('//input[@ng-reflect-name="labelFirstName"]').send_keys(firstname),
        self.chrome.find_element_by_xpath('//input[@ng-reflect-name="labelPhone"]').send_keys(phone),
        self.chrome.find_element_by_xpath('//input[@ng-reflect-name="labelCompanyName"]').send_keys(company),
        self.chrome.find_element_by_xpath('//input[@ng-reflect-name="labelRole"]').send_keys(rolecompany),
        self.chrome.find_element_by_xpath('/html/body/app-root/div[2]/app-rpa1/div/div[2]/form/input').click()])


class Excel():
    def __init__(self):
        self.arquivo = 'challenge.xlsx'
        self.arquivo_excel = load_workbook(self.arquivo)

    def pega_planilha_list(self):
        planilha = self.arquivo_excel.active
        linhas = planilha.max_row
        colunas = planilha.max_column
        # print(linhas)
        # print(colunas)
        linha_dados = []
        dados = []
        for linha in range(1,int(linhas)+1):
            linha_dados.clear()
            if linha == 1:
                continue
            else:
                for coluna in range(1,8):
                    linha_dados.append(planilha.cell(row=linha,column=coluna).value)
                dados.append(linha_dados.copy())

        return dados


if __name__ == '__main__':
    chrome = ChromeAuto()
    chrome.acessa('https://www.rpachallenge.com/?lang=EN')
    excel = Excel()
    excel.pega_planilha_list()
    lista_dados = excel.pega_planilha_list()
    chrome.inicia_contagem()
    # lastname,address,email,firstname,phone,company,rolecompany
    for linha_dados in lista_dados:
        # print(linha_dados)
        multi_process = threading.Thread(target=chrome.preenche_dados(linha_dados[1],linha_dados[4],linha_dados[5],linha_dados[0],linha_dados[6],linha_dados[2],linha_dados[3]))
        multi_process.start()
        multi_process.join()

        # p = Process(target=chrome.preenche_dados(linha_dados[1],linha_dados[4],linha_dados[5],linha_dados[0],linha_dados[6],linha_dados[2],linha_dados[3]))
        # p.start()
        # p.join()

        # chrome.preenche_dados(linha_dados[1],linha_dados[4],linha_dados[5],linha_dados[0],linha_dados[6],linha_dados[2],linha_dados[3])

    # firefox.sair()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
